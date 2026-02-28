#!/usr/bin/env python3
"""
AgentMap Excel Template Builder

Creates a multi-sheet Excel workbook for designing AgentMap workflows.
Features:
- Agent Definition tab for defining custom agents
- Prompts tab for reusable prompts
- Workflow tab with dropdowns for agent selection
- Builtin Agents reference sheet
- CSV Export sheet ready for export
- Help/Instructions sheet

Author: AgentMap Team
"""

import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.comments import Comment


# ============================================================================
# BUILTIN AGENTS DATA
# ============================================================================

BUILTIN_AGENTS = [
    # Core Agents
    {
        "Agent_ID": "builtin_echo",
        "AgentType": "echo",
        "Category": "Core",
        "Description": "Pass through and format data between nodes",
        "Input_Fields": "message",
        "Output_Field": "message",
        "Default_Prompt": "",
        "Default_Context": "{}",
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_default",
        "AgentType": "default",
        "Category": "Core",
        "Description": "Default agent that logs and returns message",
        "Input_Fields": "message",
        "Output_Field": "response",
        "Default_Prompt": "",
        "Default_Context": "{}",
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_input",
        "AgentType": "input",
        "Category": "Core",
        "Description": "Prompt user for input",
        "Input_Fields": "prompt",
        "Output_Field": "user_input",
        "Default_Prompt": "Please provide input:",
        "Default_Context": "{}",
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_branching",
        "AgentType": "branching",
        "Category": "Core",
        "Description": "Route based on field conditions",
        "Input_Fields": "status",
        "Output_Field": "result",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "success_field": "status",
            "success_values": ["ok", "success", "true"],
            "failure_values": ["error", "failed", "false"]
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_success",
        "AgentType": "success",
        "Category": "Core",
        "Description": "Always returns success (for testing)",
        "Input_Fields": "message",
        "Output_Field": "result",
        "Default_Prompt": "",
        "Default_Context": "{}",
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_failure",
        "AgentType": "failure",
        "Category": "Core",
        "Description": "Always returns failure (for testing)",
        "Input_Fields": "message",
        "Output_Field": "result",
        "Default_Prompt": "",
        "Default_Context": "{}",
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_suspend",
        "AgentType": "suspend",
        "Category": "Core",
        "Description": "Pause for external processing",
        "Input_Fields": "state",
        "Output_Field": "resumed_state",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "suspension_reason": "Awaiting external process"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_human_approval",
        "AgentType": "human",
        "Category": "Core",
        "Description": "Human approval interaction",
        "Input_Fields": "request",
        "Output_Field": "approval",
        "Default_Prompt": "Please approve the following:",
        "Default_Context": json.dumps({
            "interaction_type": "approval",
            "timeout_seconds": 3600
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_human_choice",
        "AgentType": "human",
        "Category": "Core",
        "Description": "Human choice from options",
        "Input_Fields": "options",
        "Output_Field": "selection",
        "Default_Prompt": "Please select an option:",
        "Default_Context": json.dumps({
            "interaction_type": "choice",
            "options": ["Option A", "Option B", "Option C"],
            "timeout_seconds": 3600
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_human_text",
        "AgentType": "human",
        "Category": "Core",
        "Description": "Human text input",
        "Input_Fields": "prompt",
        "Output_Field": "response",
        "Default_Prompt": "Please provide your input:",
        "Default_Context": json.dumps({
            "interaction_type": "text_input",
            "timeout_seconds": 3600
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_human_edit",
        "AgentType": "human",
        "Category": "Core",
        "Description": "Human edit existing content",
        "Input_Fields": "content",
        "Output_Field": "edited_content",
        "Default_Prompt": "Please review and edit:",
        "Default_Context": json.dumps({
            "interaction_type": "edit",
            "timeout_seconds": 3600
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_graph",
        "AgentType": "graph",
        "Category": "Core",
        "Description": "Execute a subgraph",
        "Input_Fields": "input_data",
        "Output_Field": "output_data",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "subgraph_name": "SubWorkflowName"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    # LLM Agents
    {
        "Agent_ID": "builtin_claude",
        "AgentType": "claude",
        "Category": "LLM",
        "Description": "Anthropic Claude LLM agent",
        "Input_Fields": "message",
        "Output_Field": "response",
        "Default_Prompt": "You are a helpful assistant.",
        "Default_Context": json.dumps({
            "provider": "anthropic",
            "model": "claude-sonnet-4-6",
            "temperature": 0.7,
            "max_tokens": 1000
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_gpt",
        "AgentType": "gpt",
        "Category": "LLM",
        "Description": "OpenAI GPT LLM agent",
        "Input_Fields": "message",
        "Output_Field": "response",
        "Default_Prompt": "You are a helpful assistant.",
        "Default_Context": json.dumps({
            "provider": "openai",
            "model": "gpt-4",
            "temperature": 0.7,
            "max_tokens": 1000
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_gemini",
        "AgentType": "gemini",
        "Category": "LLM",
        "Description": "Google Gemini LLM agent",
        "Input_Fields": "message",
        "Output_Field": "response",
        "Default_Prompt": "You are a helpful assistant.",
        "Default_Context": json.dumps({
            "provider": "google",
            "model": "gemini-pro",
            "temperature": 0.7,
            "max_tokens": 1000
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_llm_structured",
        "AgentType": "llm",
        "Category": "LLM",
        "Description": "Generic LLM with structured output",
        "Input_Fields": "message|context",
        "Output_Field": "structured_response",
        "Default_Prompt": "Process the input and return structured data.",
        "Default_Context": json.dumps({
            "provider": "anthropic",
            "model": "claude-sonnet-4-6",
            "temperature": 0.0,
            "max_tokens": 2000,
            "output_format": "json"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    # Storage Agents
    {
        "Agent_ID": "builtin_csv_reader",
        "AgentType": "csv_reader",
        "Category": "Storage",
        "Description": "Read data from CSV file",
        "Input_Fields": "file_path",
        "Output_Field": "data",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "encoding": "utf-8",
            "delimiter": ","
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_csv_writer",
        "AgentType": "csv_writer",
        "Category": "Storage",
        "Description": "Write data to CSV file",
        "Input_Fields": "data|file_path",
        "Output_Field": "status",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "encoding": "utf-8",
            "delimiter": ",",
            "mode": "write"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_json_reader",
        "AgentType": "json_reader",
        "Category": "Storage",
        "Description": "Read data from JSON file",
        "Input_Fields": "file_path",
        "Output_Field": "data",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "encoding": "utf-8"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_json_writer",
        "AgentType": "json_writer",
        "Category": "Storage",
        "Description": "Write data to JSON file",
        "Input_Fields": "data|file_path",
        "Output_Field": "status",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "encoding": "utf-8",
            "indent": 2
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_file_reader",
        "AgentType": "file_reader",
        "Category": "Storage",
        "Description": "Read text from file",
        "Input_Fields": "file_path",
        "Output_Field": "content",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "encoding": "utf-8"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_file_writer",
        "AgentType": "file_writer",
        "Category": "Storage",
        "Description": "Write text to file",
        "Input_Fields": "content|file_path",
        "Output_Field": "status",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "encoding": "utf-8",
            "mode": "write"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_vector_reader",
        "AgentType": "vector_reader",
        "Category": "Storage",
        "Description": "Query vector database",
        "Input_Fields": "query|top_k",
        "Output_Field": "results",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "collection": "default",
            "top_k": 5
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_vector_writer",
        "AgentType": "vector_writer",
        "Category": "Storage",
        "Description": "Write to vector database",
        "Input_Fields": "documents|embeddings",
        "Output_Field": "status",
        "Default_Prompt": "",
        "Default_Context": json.dumps({
            "collection": "default"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    # Mixed/Special Agents
    {
        "Agent_ID": "builtin_summary",
        "AgentType": "summary",
        "Category": "Mixed",
        "Description": "Summarize execution results",
        "Input_Fields": "execution_data",
        "Output_Field": "summary",
        "Default_Prompt": "Summarize the execution results.",
        "Default_Context": json.dumps({
            "format": "markdown",
            "include_metrics": True
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_orchestrator",
        "AgentType": "orchestrator",
        "Category": "Mixed",
        "Description": "Route to best node based on input",
        "Input_Fields": "input|options",
        "Output_Field": "selected_node",
        "Default_Prompt": "Select the best option for the given input.",
        "Default_Context": json.dumps({
            "selection_strategy": "llm_based"
        }),
        "Tool_Source": "",
        "Available_Tools": ""
    },
    {
        "Agent_ID": "builtin_tool_agent",
        "AgentType": "tool_agent",
        "Category": "Mixed",
        "Description": "Select and execute tools",
        "Input_Fields": "task",
        "Output_Field": "result",
        "Default_Prompt": "Execute the appropriate tool for the task.",
        "Default_Context": json.dumps({
            "max_iterations": 5
        }),
        "Tool_Source": "agentmap.tools",
        "Available_Tools": "search|calculate|fetch"
    },
]

# All unique agent types for dropdown
AGENT_TYPES = [
    "echo", "default", "input", "branching", "success", "failure",
    "suspend", "human", "graph",
    "llm", "anthropic", "claude", "openai", "gpt", "chatgpt", "google", "gemini",
    "csv_reader", "csv_writer", "json_reader", "json_writer",
    "file_reader", "file_writer", "vector_reader", "vector_writer",
    "summary", "orchestrator", "tool_agent"
]


# ============================================================================
# STYLES
# ============================================================================

def create_styles():
    """Create named styles for the workbook."""
    styles = {}

    # Header style
    styles['header'] = NamedStyle(name='header')
    styles['header'].font = Font(bold=True, color='FFFFFF', size=11)
    styles['header'].fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    styles['header'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    styles['header'].border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Subheader style
    styles['subheader'] = NamedStyle(name='subheader')
    styles['subheader'].font = Font(bold=True, color='FFFFFF', size=10)
    styles['subheader'].fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
    styles['subheader'].alignment = Alignment(horizontal='center', vertical='center')

    # Data cell style
    styles['data'] = NamedStyle(name='data')
    styles['data'].alignment = Alignment(vertical='center', wrap_text=True)
    styles['data'].border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Required field style
    styles['required'] = NamedStyle(name='required')
    styles['required'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    styles['required'].alignment = Alignment(vertical='center', wrap_text=True)
    styles['required'].border = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Title style
    styles['title'] = NamedStyle(name='title')
    styles['title'].font = Font(bold=True, size=16, color='2E75B6')
    styles['title'].alignment = Alignment(horizontal='left', vertical='center')

    return styles


# ============================================================================
# SHEET BUILDERS
# ============================================================================

def build_agent_definitions_sheet(wb, styles):
    """Build the Agent Definitions sheet."""
    ws = wb.create_sheet("Agent_Definitions")

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "Agent Definitions"
    ws['A1'].style = styles['title']
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A2:J2')
    ws['A2'] = "Define your custom agents here. These will be available in the Workflow tab dropdown."
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ("Agent_ID", "Unique identifier for this agent (used in Workflow)", 15),
        ("AgentType", "Type of agent (select from dropdown)", 15),
        ("Category", "Category for organization", 12),
        ("Description", "Description of what this agent does", 30),
        ("Input_Fields", "Input fields (pipe-separated: field1|field2)", 20),
        ("Output_Field", "Output field name", 15),
        ("Default_Prompt", "Default prompt/instructions", 40),
        ("Default_Context", "Default context JSON configuration", 40),
        ("Tool_Source", "Tool source module (for tool_agent)", 20),
        ("Available_Tools", "Available tools (for tool_agent)", 20),
    ]

    for col, (header, tooltip, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = styles['header']
        cell.comment = Comment(tooltip, "AgentMap")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 25

    # Add data validation for AgentType column
    agent_type_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(AGENT_TYPES) + '"',
        allow_blank=True
    )
    agent_type_validation.error = 'Please select a valid agent type'
    agent_type_validation.errorTitle = 'Invalid Agent Type'
    ws.add_data_validation(agent_type_validation)
    agent_type_validation.add('B4:B100')

    # Category validation
    category_validation = DataValidation(
        type="list",
        formula1='"Core,LLM,Storage,Mixed,Custom"',
        allow_blank=True
    )
    ws.add_data_validation(category_validation)
    category_validation.add('C4:C100')

    # Add example row
    example_data = [
        "my_custom_llm",
        "claude",
        "Custom",
        "My custom Claude agent for specific tasks",
        "query|context",
        "answer",
        "You are an expert assistant. Answer the query based on the context provided.",
        '{"provider": "anthropic", "model": "claude-sonnet-4-6", "temperature": 0.5}',
        "",
        ""
    ]

    for col, value in enumerate(example_data, 1):
        cell = ws.cell(row=4, column=col, value=value)
        cell.style = styles['data']

    # Pre-populate remaining rows with formatting
    for row in range(5, 50):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col, value="")
            cell.style = styles['data']

    # Freeze header row
    ws.freeze_panes = 'A4'

    return ws


def build_prompts_sheet(wb, styles):
    """Build the Prompts sheet for reusable prompts."""
    ws = wb.create_sheet("Prompts")

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "Reusable Prompts"
    ws['A1'].style = styles['title']
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A2:D2')
    ws['A2'] = "Define reusable prompts here. Reference them in the Workflow tab using @Prompt_ID"
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ("Prompt_ID", "Unique identifier (use @Prompt_ID in Workflow)", 20),
        ("Prompt_Name", "Descriptive name for the prompt", 25),
        ("Prompt_Text", "The actual prompt text", 80),
        ("Description", "Description/notes about this prompt", 40),
    ]

    for col, (header, tooltip, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = styles['header']
        cell.comment = Comment(tooltip, "AgentMap")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 25

    # Example prompts
    example_prompts = [
        (
            "system_helpful",
            "Helpful Assistant",
            "You are a helpful, harmless, and honest AI assistant. Provide clear, accurate, and useful information.",
            "General purpose system prompt"
        ),
        (
            "system_expert",
            "Domain Expert",
            "You are an expert in {domain}. Provide detailed, accurate, and professional responses based on your expertise.",
            "Expert persona prompt with domain placeholder"
        ),
        (
            "summarize_text",
            "Summarize Text",
            "Summarize the following text in a clear and concise manner. Focus on the key points and main ideas:\n\n{text}",
            "Text summarization prompt"
        ),
        (
            "extract_json",
            "Extract to JSON",
            "Extract the relevant information from the following text and return it as valid JSON:\n\n{text}\n\nReturn only the JSON, no explanation.",
            "Structured data extraction"
        ),
        (
            "analyze_sentiment",
            "Sentiment Analysis",
            "Analyze the sentiment of the following text. Classify it as positive, negative, or neutral, and explain why:\n\n{text}",
            "Sentiment analysis prompt"
        ),
    ]

    for row, (prompt_id, name, text, desc) in enumerate(example_prompts, 4):
        ws.cell(row=row, column=1, value=prompt_id).style = styles['data']
        ws.cell(row=row, column=2, value=name).style = styles['data']
        ws.cell(row=row, column=3, value=text).style = styles['data']
        ws.cell(row=row, column=4, value=desc).style = styles['data']

    # Pre-populate remaining rows
    for row in range(4 + len(example_prompts), 50):
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col, value="")
            cell.style = styles['data']

    # Freeze header row
    ws.freeze_panes = 'A4'

    return ws


def build_workflow_sheet(wb, styles):
    """Build the main Workflow sheet."""
    ws = wb.create_sheet("Workflow")

    # Title
    ws.merge_cells('A1:M1')
    ws['A1'] = "Workflow Definition"
    ws['A1'].style = styles['title']
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A2:M2')
    ws['A2'] = "Define your workflow here. Select agents from dropdown or use builtin_* agents. The CSV_Export sheet will generate the final CSV."
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws.row_dimensions[2].height = 20

    # Headers matching CSV format
    headers = [
        ("GraphName", "Name of this workflow (required)", 15),
        ("Node", "Unique node name in workflow (required)", 15),
        ("Agent_ID", "Select from Agent_Definitions or Builtin", 18),
        ("AgentType", "Agent type (auto-filled or override)", 12),
        ("Prompt", "Prompt text or @Prompt_ID reference", 35),
        ("Description", "Node description", 25),
        ("Context", "Context JSON (auto-filled or override)", 35),
        ("Input_Fields", "Input fields (auto-filled or override)", 18),
        ("Output_Field", "Output field (auto-filled or override)", 15),
        ("Edge", "Next node(s) - use | for parallel", 15),
        ("Success_Next", "Next node on success", 12),
        ("Failure_Next", "Next node on failure", 12),
        ("Notes", "Your notes (not exported)", 20),
    ]

    for col, (header, tooltip, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = styles['header']
        cell.comment = Comment(tooltip, "AgentMap")
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 25

    # Add data validation for Agent_ID (will reference Agent_Definitions and Builtin)
    # Create a named range for agent IDs (will be populated dynamically)

    # For now, add builtin agents to dropdown
    builtin_ids = [agent['Agent_ID'] for agent in BUILTIN_AGENTS]
    agent_validation = DataValidation(
        type="list",
        formula1='INDIRECT("AgentList")',  # Named range
        allow_blank=True
    )
    agent_validation.error = 'Select from Agent_Definitions or use builtin agents'
    agent_validation.errorTitle = 'Invalid Agent'
    agent_validation.prompt = 'Select an agent or type custom Agent_ID'
    agent_validation.promptTitle = 'Agent Selection'
    ws.add_data_validation(agent_validation)
    agent_validation.add('C4:C200')

    # Agent type validation for overrides
    type_validation = DataValidation(
        type="list",
        formula1='"' + ','.join(AGENT_TYPES) + '"',
        allow_blank=True
    )
    ws.add_data_validation(type_validation)
    type_validation.add('D4:D200')

    # Example workflow
    example_workflow = [
        {
            "GraphName": "ExampleWorkflow",
            "Node": "Start",
            "Agent_ID": "builtin_input",
            "AgentType": "",
            "Prompt": "What would you like help with today?",
            "Description": "Get initial user input",
            "Context": "",
            "Input_Fields": "",
            "Output_Field": "",
            "Edge": "Process",
            "Success_Next": "",
            "Failure_Next": "",
            "Notes": "Entry point"
        },
        {
            "GraphName": "ExampleWorkflow",
            "Node": "Process",
            "Agent_ID": "builtin_claude",
            "AgentType": "",
            "Prompt": "@system_helpful",
            "Description": "Process user request",
            "Context": "",
            "Input_Fields": "",
            "Output_Field": "",
            "Edge": "",
            "Success_Next": "Respond",
            "Failure_Next": "HandleError",
            "Notes": "Main processing"
        },
        {
            "GraphName": "ExampleWorkflow",
            "Node": "Respond",
            "Agent_ID": "builtin_echo",
            "AgentType": "",
            "Prompt": "",
            "Description": "Return response to user",
            "Context": "",
            "Input_Fields": "",
            "Output_Field": "",
            "Edge": "",
            "Success_Next": "",
            "Failure_Next": "",
            "Notes": "Output"
        },
        {
            "GraphName": "ExampleWorkflow",
            "Node": "HandleError",
            "Agent_ID": "builtin_default",
            "AgentType": "",
            "Prompt": "An error occurred. Please try again.",
            "Description": "Handle errors",
            "Context": "",
            "Input_Fields": "",
            "Output_Field": "",
            "Edge": "",
            "Success_Next": "",
            "Failure_Next": "",
            "Notes": "Error handler"
        },
    ]

    for row_idx, data in enumerate(example_workflow, 4):
        for col_idx, key in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=data.get(key[0], ""))
            if col_idx <= 2:  # Required fields
                cell.style = styles['required']
            else:
                cell.style = styles['data']

    # Pre-populate remaining rows
    for row in range(4 + len(example_workflow), 200):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col, value="")
            if col <= 2:
                cell.style = styles['required']
            else:
                cell.style = styles['data']

    # Freeze header row
    ws.freeze_panes = 'A4'

    return ws


def build_builtin_agents_sheet(wb, styles):
    """Build the Builtin Agents reference sheet."""
    ws = wb.create_sheet("Builtin_Agents")

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = "Builtin Agents Reference"
    ws['A1'].style = styles['title']
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A2:J2')
    ws['A2'] = "Reference sheet for all builtin agents. These can be selected directly in the Workflow tab."
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws.row_dimensions[2].height = 20

    # Headers
    headers = [
        ("Agent_ID", 20),
        ("AgentType", 12),
        ("Category", 10),
        ("Description", 35),
        ("Input_Fields", 20),
        ("Output_Field", 15),
        ("Default_Prompt", 35),
        ("Default_Context", 40),
        ("Tool_Source", 18),
        ("Available_Tools", 18),
    ]

    for col, (header, width) in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = styles['header']
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 25

    # Populate builtin agents
    for row_idx, agent in enumerate(BUILTIN_AGENTS, 4):
        for col_idx, (key, _) in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=agent.get(key, ""))
            cell.style = styles['data']

    # Color code by category
    category_colors = {
        "Core": "E2EFDA",    # Green
        "LLM": "DDEBF7",     # Blue
        "Storage": "FCE4D6",  # Orange
        "Mixed": "E4DFEC",    # Purple
    }

    for row_idx, agent in enumerate(BUILTIN_AGENTS, 4):
        category = agent.get("Category", "")
        if category in category_colors:
            for col in range(1, 11):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=category_colors[category],
                    end_color=category_colors[category],
                    fill_type='solid'
                )

    # Freeze header row
    ws.freeze_panes = 'A4'

    # Protect sheet (reference only)
    ws.protection.sheet = True
    ws.protection.password = ''  # No password, just prevent accidental edits

    return ws


def build_csv_export_sheet(wb, styles):
    """Build the CSV Export sheet with formulas to pull from Workflow."""
    ws = wb.create_sheet("CSV_Export")

    # Title
    ws.merge_cells('A1:L1')
    ws['A1'] = "CSV Export (Copy this sheet to export as CSV)"
    ws['A1'].style = styles['title']
    ws.row_dimensions[1].height = 30

    # Instructions
    ws.merge_cells('A2:L2')
    ws['A2'] = "This sheet shows the final CSV output. Copy all data (Ctrl+C) and paste into a text file, or use File > Save As > CSV."
    ws['A2'].font = Font(italic=True, size=10, color='666666')
    ws.row_dimensions[2].height = 20

    # CSV Headers (exact AgentMap format)
    csv_headers = [
        ("GraphName", 15),
        ("Node", 15),
        ("AgentType", 12),
        ("Prompt", 40),
        ("Description", 30),
        ("Context", 40),
        ("Input_Fields", 20),
        ("Output_Field", 15),
        ("Edge", 15),
        ("Success_Next", 12),
        ("Failure_Next", 12),
    ]

    for col, (header, width) in enumerate(csv_headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.style = styles['header']
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[3].height = 25

    # Add formulas to pull from Workflow sheet with agent resolution
    # Map columns from Workflow to CSV Export
    # Workflow columns: A=GraphName, B=Node, C=Agent_ID, D=AgentType, E=Prompt,
    #                   F=Description, G=Context, H=Input_Fields, I=Output_Field,
    #                   J=Edge, K=Success_Next, L=Failure_Next

    for row in range(4, 200):
        workflow_row = row

        # GraphName - direct copy
        ws.cell(row=row, column=1, value=f"=IF(Workflow!A{workflow_row}=\"\",\"\",Workflow!A{workflow_row})")

        # Node - direct copy
        ws.cell(row=row, column=2, value=f"=IF(Workflow!B{workflow_row}=\"\",\"\",Workflow!B{workflow_row})")

        # AgentType - use override if provided, else lookup from Agent_ID
        ws.cell(row=row, column=3, value=f'=IF(Workflow!B{workflow_row}="","",IF(Workflow!D{workflow_row}<>"",Workflow!D{workflow_row},IFERROR(VLOOKUP(Workflow!C{workflow_row},Builtin_Agents!A:B,2,FALSE),IFERROR(VLOOKUP(Workflow!C{workflow_row},Agent_Definitions!A:B,2,FALSE),""))))')

        # Prompt - resolve @references or use direct value
        ws.cell(row=row, column=4, value=f'=IF(Workflow!B{workflow_row}="","",IF(LEFT(Workflow!E{workflow_row},1)="@",IFERROR(VLOOKUP(MID(Workflow!E{workflow_row},2,100),Prompts!A:C,3,FALSE),Workflow!E{workflow_row}),Workflow!E{workflow_row}))')

        # Description - direct copy
        ws.cell(row=row, column=5, value=f"=IF(Workflow!B{workflow_row}=\"\",\"\",Workflow!F{workflow_row})")

        # Context - use override if provided, else lookup
        ws.cell(row=row, column=6, value=f'=IF(Workflow!B{workflow_row}="","",IF(Workflow!G{workflow_row}<>"",Workflow!G{workflow_row},IFERROR(VLOOKUP(Workflow!C{workflow_row},Builtin_Agents!A:H,8,FALSE),IFERROR(VLOOKUP(Workflow!C{workflow_row},Agent_Definitions!A:H,8,FALSE),""))))')

        # Input_Fields - use override if provided, else lookup
        ws.cell(row=row, column=7, value=f'=IF(Workflow!B{workflow_row}="","",IF(Workflow!H{workflow_row}<>"",Workflow!H{workflow_row},IFERROR(VLOOKUP(Workflow!C{workflow_row},Builtin_Agents!A:E,5,FALSE),IFERROR(VLOOKUP(Workflow!C{workflow_row},Agent_Definitions!A:E,5,FALSE),""))))')

        # Output_Field - use override if provided, else lookup
        ws.cell(row=row, column=8, value=f'=IF(Workflow!B{workflow_row}="","",IF(Workflow!I{workflow_row}<>"",Workflow!I{workflow_row},IFERROR(VLOOKUP(Workflow!C{workflow_row},Builtin_Agents!A:F,6,FALSE),IFERROR(VLOOKUP(Workflow!C{workflow_row},Agent_Definitions!A:F,6,FALSE),""))))')

        # Edge - direct copy
        ws.cell(row=row, column=9, value=f"=IF(Workflow!B{workflow_row}=\"\",\"\",Workflow!J{workflow_row})")

        # Success_Next - direct copy
        ws.cell(row=row, column=10, value=f"=IF(Workflow!B{workflow_row}=\"\",\"\",Workflow!K{workflow_row})")

        # Failure_Next - direct copy
        ws.cell(row=row, column=11, value=f"=IF(Workflow!B{workflow_row}=\"\",\"\",Workflow!L{workflow_row})")

        # Apply styling
        for col in range(1, 12):
            ws.cell(row=row, column=col).style = styles['data']

    # Freeze header row
    ws.freeze_panes = 'A4'

    return ws


def build_help_sheet(wb, styles):
    """Build the Help/Instructions sheet."""
    ws = wb.create_sheet("Help")

    # Title
    ws.merge_cells('A1:D1')
    ws['A1'] = "AgentMap Workflow Builder - Help & Instructions"
    ws['A1'].style = styles['title']
    ws.row_dimensions[1].height = 35

    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 5

    help_content = [
        ("", "", ""),
        ("OVERVIEW", "", ""),
        ("", "This Excel workbook helps you design AgentMap workflows visually.", ""),
        ("", "Define agents, create prompts, build workflows, and export to CSV.", ""),
        ("", "", ""),
        ("SHEETS", "", ""),
        ("", "1. Agent_Definitions", "Define your custom agents with their configurations"),
        ("", "2. Prompts", "Create reusable prompts that can be referenced anywhere"),
        ("", "3. Workflow", "Build your workflow by selecting agents and defining flow"),
        ("", "4. Builtin_Agents", "Reference sheet of all available builtin agents"),
        ("", "5. CSV_Export", "Final output - copy this to export as CSV"),
        ("", "", ""),
        ("HOW TO USE", "", ""),
        ("", "Step 1:", "Define any custom agents in Agent_Definitions tab"),
        ("", "Step 2:", "Create reusable prompts in Prompts tab (optional)"),
        ("", "Step 3:", "Build workflow in Workflow tab:"),
        ("", "  - Set GraphName", "(same for all nodes in a workflow)"),
        ("", "  - Add Node names", "(unique within the workflow)"),
        ("", "  - Select Agent_ID", "(from dropdown or type custom)"),
        ("", "  - Set prompts", "(direct text or @Prompt_ID reference)"),
        ("", "  - Define routing", "(Edge, Success_Next, or Failure_Next)"),
        ("", "Step 4:", "Review CSV_Export tab"),
        ("", "Step 5:", "Export: Select all in CSV_Export, copy, paste to .csv file"),
        ("", "", ""),
        ("ROUTING RULES", "", ""),
        ("", "Simple flow:", "Use Edge column (e.g., 'NextNode')"),
        ("", "Conditional flow:", "Use Success_Next and Failure_Next"),
        ("", "Parallel flow:", "Use pipe separator (e.g., 'Node1|Node2|Node3')"),
        ("", "Important:", "Don't mix Edge with Success_Next/Failure_Next"),
        ("", "", ""),
        ("FIELD FORMAT", "", ""),
        ("", "Input_Fields:", "Pipe-separated (e.g., 'query|context|options')"),
        ("", "Context:", "Valid JSON (e.g., {\"temperature\": 0.7})"),
        ("", "Prompt refs:", "Use @Prompt_ID to reference Prompts sheet"),
        ("", "", ""),
        ("AGENT TYPES", "", ""),
        ("", "Core:", "echo, default, input, branching, human, graph"),
        ("", "LLM:", "claude, gpt, gemini, llm (generic)"),
        ("", "Storage:", "csv_reader/writer, json_reader/writer, file_reader/writer"),
        ("", "Mixed:", "summary, orchestrator, tool_agent"),
        ("", "", ""),
        ("TIPS", "", ""),
        ("", "- Use builtin agents", "for common tasks (already configured)"),
        ("", "- Auto-fill works", "Input/Output/Context auto-populate from agent"),
        ("", "- Override as needed", "Fill in fields to override agent defaults"),
        ("", "- Notes column", "Use it! Not exported, just for your reference"),
        ("", "", ""),
        ("EXPORT INSTRUCTIONS", "", ""),
        ("", "1.", "Go to CSV_Export sheet"),
        ("", "2.", "Select from A3 to last row with data"),
        ("", "3.", "Copy (Ctrl+C)"),
        ("", "4.", "Open text editor, paste, save as .csv"),
        ("", "Or:", "File > Save As > CSV (may need to select CSV_Export sheet first)"),
    ]

    for row_idx, (section, label, content) in enumerate(help_content, 2):
        if section:
            ws.cell(row=row_idx, column=2, value=section)
            ws.cell(row=row_idx, column=2).font = Font(bold=True, size=12, color='2E75B6')
        if label:
            ws.cell(row=row_idx, column=2, value=label)
            if section == "":
                ws.cell(row=row_idx, column=2).font = Font(bold=True)
        if content:
            ws.cell(row=row_idx, column=3, value=content)

    return ws


def create_named_ranges(wb):
    """Create named ranges for data validation and lookups."""
    # Get sheet references
    agent_def_sheet = wb["Agent_Definitions"]
    builtin_sheet = wb["Builtin_Agents"]

    # Create a list of all agent IDs (builtin + custom)
    # This will be used as a dropdown source

    # For simplicity, we'll create a helper column in a hidden sheet
    # with all agent IDs combined

    helper_sheet = wb.create_sheet("_Helper")
    helper_sheet.sheet_state = 'hidden'

    # Add builtin agent IDs
    row = 1
    for agent in BUILTIN_AGENTS:
        helper_sheet.cell(row=row, column=1, value=agent['Agent_ID'])
        row += 1

    # Add formula to pull custom agent IDs
    for i in range(50):  # Support up to 50 custom agents
        helper_sheet.cell(row=row + i, column=1,
                         value=f"=IF(Agent_Definitions!A{4+i}=\"\",\"\",Agent_Definitions!A{4+i})")

    # Create named range for agent list
    from openpyxl.workbook.defined_name import DefinedName

    # Count builtin agents + custom agent slots
    total_rows = len(BUILTIN_AGENTS) + 50
    ref = f"'_Helper'!$A$1:$A${total_rows}"
    defn = DefinedName("AgentList", attr_text=ref)
    wb.defined_names.add(defn)


def build_agentmap_template():
    """Main function to build the complete Excel template."""
    print("Building AgentMap Workflow Template...")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create styles
    styles = create_styles()

    # Register styles
    for style in styles.values():
        try:
            wb.add_named_style(style)
        except ValueError:
            pass  # Style already exists

    # Build sheets
    print("  Creating Agent_Definitions sheet...")
    build_agent_definitions_sheet(wb, styles)

    print("  Creating Prompts sheet...")
    build_prompts_sheet(wb, styles)

    print("  Creating Workflow sheet...")
    build_workflow_sheet(wb, styles)

    print("  Creating Builtin_Agents reference sheet...")
    build_builtin_agents_sheet(wb, styles)

    print("  Creating CSV_Export sheet...")
    build_csv_export_sheet(wb, styles)

    print("  Creating Help sheet...")
    build_help_sheet(wb, styles)

    print("  Creating named ranges...")
    create_named_ranges(wb)

    # Set Workflow as the active sheet
    wb.active = wb["Workflow"]

    # Save workbook
    output_path = "/home/user/AgentMap/templates/AgentMap_Workflow_Template.xlsx"

    # Ensure templates directory exists
    import os
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    wb.save(output_path)
    print(f"\nTemplate saved to: {output_path}")

    return output_path


if __name__ == "__main__":
    build_agentmap_template()
